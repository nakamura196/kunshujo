import urllib.request
from bs4 import BeautifulSoup
import csv
from time import sleep
import pandas as pd
import json
import urllib.request
import os
from PIL import Image
import yaml
import requests
import sys
import argparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import openpyxl

hyperlink = Font(underline='single', color='0563C1')

def createExcel(headers, members, mid):


    wb = openpyxl.Workbook()
    ws = wb.active

    for c in range(len(headers)):
        ws.cell(row=1, column=1 + c).value = headers[c]
        # ws.cell(row=1, column=1).value = '=HYPERLINK("https://static.toyobunko-lab.jp/taishozo/search/?keyword={}", "{}")'.format("ccc", "bbb")
        # ws.cell(row=1, column=1).font = hyperlink

    

    for r in range(len(members)):
        member = members[r]

        for c in range(len(headers)):
            header = headers[c]
            row = 2 + r
            col = 1 + c
            value = member[header]
            if header in ["source"]:
                ws.cell(row=row, column=col).value = '=HYPERLINK("{}", "{}")'.format(value, value)
                ws.cell(row=row, column=col).font = hyperlink
            else:
                ws.cell(row=row, column=col).value = value

    x_path = "../static/data/metadata/{}.xlsx".format(mid)
    wb.save(x_path)

def createCsv(headers, members, mid):
    rows = []
    row = []
    rows.append(row)
    for c in range(len(headers)):
        row.append(headers[c])

    for r in range(len(members)):
        member = members[r]

        row = []
        rows.append(row)

        for c in range(len(headers)):
            header = headers[c]
            value = member[header]
            row.append(value)

    x_path = "../static/data/metadata/{}.xlsx".format(mid)
    df = pd.DataFrame(rows)
    df.to_csv(x_path.replace(".xlsx", ".csv"), header=False, index=False)


f = open("../static/data/index.json", 'r')
index = json.load(f)

map = {}

for item in index:
    id = item["objectID"]
    print(id)

    manifest = item["manifest"]

    pic = item["図"][0]

    if manifest not in map:
        map[manifest] = {
            "members": {},
            "label" : pic, # row["parent:label"],
            # "uri": row["schema:isPartOf^^uri"]
        }

    members = map[manifest]["members"]

    category = item["category"][0]

    marker = "https://www.hi.u-tokyo.ac.jp/collection/degitalgallary/ryukyu/data/files/legend/" + category + ".png"

    member_id = item["member"]

    label = item["label"]

    prefix2 = "https://www.hi.u-tokyo.ac.jp/collection/degitalgallary/ryukyu"
    url = prefix2 + "/item/" + item["objectID"]

    metadata = [
        {
            "label": "Annotation",
            "value": [
                {
                    "@id": member_id, # url,
                    "@type": "oa:Annotation",
                    "motivation": "sc:painting",
                    "on": member_id,
                    "resource": {
                        "@type": "cnt:ContentAsText",
                        
                        "format": "text/html",
                        "marker": {
                        "@id": marker + "#xy=12,22",
                        "@type": "dctypes:Image"
                        }
                    }
                }
            ]
        }
    ]

    

    metadata.append({
        "label" : "図",
        "value" : pic
    })

    metadata.append({
        "label" : "category",
        "value" : category
    })

    member = {
        "entry_id": id,
        "body": label,
        "source": "https://www.hi.u-tokyo.ac.jp/collection/degitalgallary/ryukyu/item/{}".format(id),
        "ne_class": item["category"][0],
        "latitude": item["latitude"][0],
        "longitude": item["longitude"][0],
        "description": pic,
        
        "説明文": item["description"][0],
        # "石高": item["description"][0],
        "備考": item["備考"][0],

        "沖縄県教育委員会編『琉球国絵図史料集』第１集の番号": item["番号"][0],
        "ジャパンナレッジID": item["jk"][0],
        "現在の地名": item["現在の地名"][0],

    }

    members[id] = member

headers = [
    "entry_id",
    "body",
    "ne_class",
    "latitude",
    "longitude",
    "description",
    "説明文",
    "備考",
    "沖縄県教育委員会編『琉球国絵図史料集』第１集の番号",
    "ジャパンナレッジID",
    "現在の地名",
     "source",
]

rows2 = []
row = []
rows2.append(row)
for c in range(len(headers)):
    row.append(headers[c])

top_path = "../static/data/metadata/top.xlsx"

members2 = []

for manifest in map:
    mid = manifest.split("/")[-2]

    membersMap = map[manifest]["members"]

    members = []

    for id in sorted(membersMap):
        members.append(membersMap[id])
        members2.append(membersMap[id])

    createExcel(headers, members, mid)

    # csv

    createCsv(headers, members, mid)

    
createExcel(headers, members2,  "top")
createCsv(headers, members2,  "top")